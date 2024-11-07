import json
import re
import typing
from dataclasses import dataclass
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor

from DrissionPage import ChromiumOptions, ChromiumPage
from DrissionPage.common import Settings
from DrissionPage.errors import ElementNotFoundError, WaitTimeoutError
from rich import print

DEBUG = False
DT_FMT = "%Y-%m-%d %H:%M:%S"

Settings.raise_when_wait_failed = True


@dataclass
class Article:
    title: str
    url: str
    published: datetime
    content: str


class PageNotFound(Exception):
    pass


class Fetcher:
    def __init__(self, code: str, date: date):
        self.code = code
        self.start_url = f"https://guba.eastmoney.com/list,{code}"
        self.date = date
        self.page = ChromiumPage(self.get_chrome_options())
        self.total_page = -1

    @staticmethod
    def get_chrome_options() -> ChromiumOptions:
        co = ChromiumOptions()
        co.set_argument("--no-sandbox")
        if not DEBUG:
            co.headless()
            co.set_user_agent(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
            )
        return co

    def get_url(self, pn: int) -> str:
        return f"{self.start_url}_{pn}.html"

    def get_article_url(self, post_id: int) -> str:
        return f"https://guba.eastmoney.com/news,{self.code},{post_id}.html"

    def get_article_list(self, pn: int) -> typing.Generator[Article, None, None]:
        self.page.get(self.get_url(pn))
        print(f"[green] Fetching page {self.get_url(pn)}...")
        article_list = re.search(
            r"var article_list\s*=\s*({.*?});", self.page.html
        ).group(1)

        json_data = json.loads(article_list)
        articles = json_data["re"]
        for article in articles:
            yield Article(
                title=article["post_title"],
                url=self.get_article_url(article["post_id"]),
                published=datetime.strptime(article["post_publish_time"], DT_FMT),
                content="",
            )

    def get_total_page(self):
        if self.total_page == -1:
            print(self.get_url(1))
            self.page.get(self.get_url(1), retry=3)
            if self.page.url == "https://guba.eastmoney.com/error?type=1":
                raise PageNotFound("Page not found")
            self.page.wait.ele_displayed("t:ul@class:paging")
            pagers = self.page.ele("t:ul@class:paging").eles("t:li")
            last_page = pagers[-2].text
            self.total_page = int(last_page)
            print(f"Total page: {self.total_page}")

    def get_article_detail(self, article: Article) -> Article:
        print(f"[green] {article.published} {article.title}")
        new_tab = self.page.new_tab()
        try:
            new_tab.get(article.url)
            new_tab.wait.ele_displayed("@class:newstext")
            article.content = new_tab.ele("@class:newstext").text
            return article
        except:
            return article
        finally:
            new_tab.close()

    @staticmethod
    def get_date_range(articles: typing.Iterable[Article]) -> tuple[date, date]:
        articles = list(articles)

        return articles[0].published.date(), articles[-1].published.date()

    def get_articles_with_date(self) -> typing.Generator[Article, None, None]:
        self.get_total_page()

        # Binary search
        start, end = 1, self.total_page

        while start <= end:
            print(f"[blue] Searching: {start} - {end}")
            mid = (start + end) // 2
            earliest_date_mid, latest_date_mid = self.get_date_range(
                self.get_article_list(mid)
            )

            if self.date > earliest_date_mid:
                end = mid - 1
            elif self.date < latest_date_mid:
                start = mid + 1
            else:
                end = mid - 1

        print(f"目标页面: {start}")

        # 获取目标页前后3天的文章
        start_page = start - 1 if start - 1 > 0 else 1
        end_page = end + 1 if end + 1 <= self.total_page else self.total_page

        # 未做优化，maxworker过大会爆内存
        with ThreadPoolExecutor(max_workers=4) as executor:
            for page in range(start_page, end_page + 1):
                print(f"[blue] Fetching page {page}...")
                yield from executor.map(
                    self.get_article_detail, self.get_article_list(page)
                )


if __name__ == "__main__":
    import csv
    from openpyxl import load_workbook

    wb = load_workbook("data_china.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, dt, _ = row
        dt = dt.date()

        output = f"{code}.csv"
        print(f"Fetching {code} on {dt}...")
        fetcher = Fetcher(code, dt)

        with open(output, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["标题", "链接", "发布时间", "内容"])
            try:
                for article in fetcher.get_articles_with_date():
                    writer.writerow(
                        [article.title, article.url, article.published, article.content]
                    )
            except PageNotFound:
                print(f"Page not found: {code}")
            except ElementNotFoundError:
                print(f"Element not found: {code}")
            except WaitTimeoutError:
                print(f"Wait timeout: {code}")


        fetcher.page.close()
